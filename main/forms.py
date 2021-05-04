from io import BytesIO

from django import forms
from crispy_forms.helper import FormHelper
from openpyxl import load_workbook

from employees.models import Company, Employee
import os
from zipfile import ZipFile


class DriverFilterForm(forms.Form):
    SORT_CHOICES = [
        ('', 'Sort By'),
        ('last_name', 'Last Name'),
        ('first_name', 'First Name'),
        ('employee_id', 'Employee ID'),
        ('position', 'Position'),
        ('hire_date', 'Hire Date'),
    ]

    company = forms.CharField(initial='Company Filter', required=False)
    position = forms.CharField(initial='Position Filter', required=False)
    sort_by = forms.CharField(initial='Sort By', widget=forms.Select(choices=SORT_CHOICES,
                                                                     attrs={'style': 'font-size: 14px',
                                                                            'onchange': 'form.submit();'}),
                              required=False)
    search = forms.CharField(max_length=30, widget=forms.TextInput(attrs={'placeholder': 'Search by Name or ID',
                                                                          'style': 'font-size: 14px'}),
                             required=False)

    def __init__(self, *args, **kwargs):
        super(DriverFilterForm, self).__init__(*args, **kwargs)
        self.helper = FormHelper()
        self.helper.form_show_labels = False

        company_choices = [(company.display_name, company.display_name) for company in Company.objects.all()]
        company_choices.insert(0, ('', 'Company Filter'))
        position_choices = Employee.POSITION_CHOICES
        position_choices[0] = ('', 'Position Filter')

        self.fields['company'].widget = forms.Select(choices=company_choices, attrs={'style': 'font-size: 14px',
                                                                                     'onchange': 'form.submit();'})
        self.fields['position'].widget = forms.Select(choices=position_choices, attrs={'style': 'font-size: 14px',
                                                                                       'onchange': 'form.submit();'})


class DriverImportForm(forms.Form):
    driver_file_import = forms.FileField(label='File', help_text='Ensure this is a .zip', required=True)

    def clean(self):
        super().clean()

        field_name = 'driver_file_import'
        file = self.files[field_name]
        file_name, file_extension = os.path.splitext(file.name)

        if file_extension != '.zip':
            self.add_error(field_name, f'File must be a zipped folder')
        else:
            with ZipFile(self.files['driver_file_import'], 'r') as zipfile:
                try:
                    driver_info = zipfile.read('drivers/drivers.xlsx')

                    wb = load_workbook(filename=BytesIO(driver_info))

                    try:
                        sheet = wb['data']

                        correct_headers = ['last_name', 'first_name', 'employee_id', 'position', 'hire_date',
                                           'application_date', 'classroom_date', 'company', 'is_partime',
                                           'primary_phone', 'secondary_phone', 'SS']
                        headers = []
                        for col in sheet.iter_cols():
                            headers.append(col[0].value)

                        if headers != correct_headers:
                            self.add_error(field_name, 'Headers are incorrect,'
                                                       ' look at documentation for more information')

                    except KeyError:
                        self.add_error(field_name, 'No sheet named "data"')
                except KeyError:
                    self.add_error(field_name, 'No excel file named "drivers.xlsx"')


class AttendanceImportForm(forms.Form):
    attendance_file_import = forms.FileField(label='File', help_text='Ensure this is a .xlsx', required=True)

    def clean(self):
        super().clean()

        field_name = 'attendance_file_import'
        file = self.cleaned_data[field_name]
        file_name, file_extension = os.path.splitext(file.name)

        if file_extension != '.xlsx':
            self.add_error(field_name, f'File must be a excel file')
        else:
            wb = load_workbook(filename=BytesIO(self.files['attendance_file_import'].read()))

            try:
                sheet = wb['data']

                correct_headers = ['employee_id', 'incident_date', 'reason', 'exemption', 'assigned_by']
                headers = []
                for col in sheet.iter_cols():
                    headers.append(col[0].value)

                if headers != correct_headers:
                    self.add_error(field_name, 'Headers are incorrect,'
                                               ' look at documentation for more information')

            except KeyError:
                self.add_error(field_name, 'No sheet named "data"')


class SafetyPointImportForm(forms.Form):
    safety_point_file_import = forms.FileField(label='File', help_text='Ensure this is a .xlsx', required=True)

    def clean(self):
        super().clean()

        field_name = 'safety_point_file_import'
        file = self.cleaned_data[field_name]
        file_name, file_extension = os.path.splitext(file.name)

        if file_extension != '.xlsx':
            self.add_error(field_name, f'File must be a excel file')
        else:
            wb = load_workbook(filename=BytesIO(self.files['safety_point_file_import'].read()))

            try:
                sheet = wb['data']

                correct_headers = ['employee_id', 'incident_date', 'issued_date', 'reason', 'assigned_by']
                headers = []
                for col in sheet.iter_cols():
                    headers.append(col[0].value)

                if headers != correct_headers:
                    self.add_error(field_name, 'Headers are incorrect,'
                                               ' look at documentation for more information')

            except KeyError:
                self.add_error(field_name, 'No sheet named "data"')
