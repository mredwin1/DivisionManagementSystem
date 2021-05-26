import datetime
import os

from io import BytesIO
from zipfile import ZipFile

from celery import shared_task
from django.core.mail import send_mail
from django.core.management import call_command
from django.contrib.sites.models import Site
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from twilio.rest import Client

from employees.models import Company, Employee, Attendance, SafetyPoint, Counseling


@shared_task
def send_email(subject, plain_message, to, html_message):
    send_mail(subject=subject, from_email=None, message=plain_message, recipient_list=[to], html_message=html_message)


@shared_task
def send_text(to, body, sender_id, sender_type):
    client = Client(os.environ['TWILIO_ACCOUNT_SID'], os.environ['TWILIO_AUTH_TOKEN'])
    domain = Site.objects.get_current().domain

    message = client.messages.create(
        to=to,
        from_=os.environ['TWILIO_PHONE_NUMBER'],
        body=body,
        status_callback=f'https://{domain}{reverse("main-update-msg-status", args=[sender_id, sender_type])}'
    )

    message_status = message.status

    record_types = {
        'Attendance': Attendance,
        'Safety Point': SafetyPoint,
        'Counseling': Counseling
    }

    record = record_types[sender_type].objects.get(id=sender_id)
    import logging
    logging.info(message_status)
    # record.message_status = f'{message_status[0].upper}{message_status[1:]}'
    # record.status_update_date = timezone.now()
    # record.save()


@shared_task
def import_drivers(path):
    with ZipFile(path) as zipfile:
        try:
            driver_info = zipfile.read('drivers/drivers.xlsx')

            wb = load_workbook(filename=BytesIO(driver_info))
            try:
                sheet = wb['data']
                counter = 0
                for row in sheet.iter_rows(min_row=2):
                    try:
                        last_name = f'{row[0].value.lower()[0].upper()}{row[0].value.lower()[1:]}'
                        first_name = f'{row[1].value.lower()[0].upper()}{row[1].value.lower()[1:]}'
                        employee_id = row[2].value
                        position = row[3].value.lower()
                        hire_date = datetime.datetime.strptime(str(row[4].value), '%Y%m%d')
                        application_date = datetime.datetime.strptime(str(row[5].value), '%Y%m%d')
                        classroom_date = datetime.datetime.strptime(str(row[6].value), '%Y%m%d')
                        company_name = row[7].value
                        is_part_time = True if row[8].value == 'TRUE' else False
                        primary_phone = row[9].value
                        secondary_phone = row[10].value
                        ss_number = row[11].value

                        company = Company.objects.get(display_name=company_name)

                        new_employee = Employee(
                            first_name=first_name,
                            last_name=last_name,
                            employee_id=employee_id,
                            primary_phone=primary_phone,
                            secondary_phone=secondary_phone,
                            hire_date=hire_date,
                            application_date=application_date,
                            classroom_date=classroom_date,
                            company=company,
                            is_part_time=is_part_time,
                            username=employee_id,
                            position=position,
                            is_active=True,
                        )

                        password = f'{company}{first_name[0].upper()}{last_name[0].upper()}{ss_number}'

                        new_employee.set_password(password)

                        try:
                            profile_picture = BytesIO(zipfile.read(f'drivers/pictures/{employee_id}.jpg'))
                            new_employee.profile_picture.save(f'{employee_id}.jpg', profile_picture)
                        except KeyError:
                            pass

                        new_employee.save()
                        counter += 1
                    except:
                        pass
            except KeyError:
                pass
        except KeyError:
            pass
    try:
        os.remove(path)
    except FileNotFoundError:
        pass

    return f'Imported a total of {counter} employees'


@shared_task
def import_attendance(path):
    with open(path, 'rb') as f:
        wb = load_workbook(filename=f)

        try:
            sheet = wb['data']

        except KeyError:
            pass
        for row in sheet.iter_rows(min_row=2):
            points = {
                '0': 1,
                '1': 0,
                '2': 1.5,
                '3': 4,
                '4': 1,
                '5': 1,
                '6': .5,
                '7': 1,
                '8': .5,
            }

            try:
                employee = Employee.objects.get(employee_id=int(row[0].value))
                incident_date = datetime.datetime.strptime(str(row[1].value), '%Y%m%d')
                reason = str(row[2].value)
                exemption = str(row[3].value) if row[3].value else ''
                assigned_by = int(row[4].value)
                point = 0 if exemption else points[reason]

                if exemption == '1':
                    employee.paid_sick -= 1
                if exemption == '2':
                    employee.unpaid_sick -= 1

                employee.save()

                new_attendance = Attendance(
                    employee=employee,
                    incident_date=incident_date,
                    reason=reason,
                    exemption=exemption,
                    assigned_by=assigned_by,
                    points=point,
                    uploaded=True
                )

                new_attendance.save()
            except Employee.DoesNotExist:
                pass
    try:
        os.remove(path)
    except FileNotFoundError:
        pass


@shared_task
def import_safety_points(path):
    with open(path, 'rb') as f:
        wb = load_workbook(filename=f)

        try:
            sheet = wb['data']

        except KeyError:
            pass

        for row in sheet.iter_rows(min_row=2):
            points = {
                '0': 1,
                '1': 1,
                '2': 1,
                '3': 2,
                '4': 2,
                '5': 2,
                '6': 2,
                '7': 3,
                '8': 4,
                '9': 6,
                '10': 6,
                '11': 6,
                '12': 6,
                '13': 6,
                '14': 6,
            }

            try:
                employee = Employee.objects.get(employee_id=int(row[0].value))
                incident_date = datetime.datetime.strptime(str(row[1].value), '%Y%m%d')
                issued_date = datetime.datetime.strptime(str(row[2].value), '%Y%m%d')
                reason = str(row[3].value)
                assigned_by = int(row[4].value)

                new_safety_point = SafetyPoint(
                    employee=employee,
                    incident_date=incident_date,
                    issued_date=issued_date,
                    reason=reason,
                    assigned_by=assigned_by,
                    points=points[reason]
                )

                new_safety_point.save()
            except Employee.DoesNotExist:
                pass
    try:
        os.remove(path)
    except FileNotFoundError:
        pass


@shared_task
def attendance_cleanup():
    call_command('attendance_cleanup')


@shared_task
def counseling_cleanup():
    call_command('counseling_cleanup')


@shared_task
def safety_point_cleanup():
    call_command('safety_point_cleanup')


@shared_task
def time_off_cleanup():
    call_command('time_off_cleanup')


@shared_task
def attendance_notification():
    call_command('attendance_notification')


@shared_task
def safety_point_notification():
    call_command('safety_point_notification')


@shared_task
def counseling_notification():
    call_command('counseling_notification')


@shared_task
def settlement_notification():
    call_command('settlement_notification')


@shared_task
def employee_check():
    call_command('sick_day_check')
    call_command('floating_holiday_check')
